import openpyxl
wk=openpyxl.load_workbook("C:\\Users\\kdruthiv\\PycharmProjects\\RobotFrameworkPrac\\ExcelSheets\\Excel1.xlsx")
sh=wk["Sheet1"]
rows=sh.max_row
columns=sh.max_column
print("Total rows are: "+str(rows))
print("Total columns are: "+str(columns))

'''
for i in range(1,rows+1):
    for j in range(1,columns+1):
        c=sh.cell(i,j)
        print(c.value) '''
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)


